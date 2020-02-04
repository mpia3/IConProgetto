# -*- coding: utf-8 -*-
"""
In this file there are functions useful to handle a KB object 
"""


from logicProblem import KB, Clause
from logicBottomUp import fixed_point
from prob_utils import ProbItem
import os 

def get_disease(kb= KB()):
    """
    Returns all diseases true in the KB fixed point, starting from the synthoms true
    """
    fp = fixed_point(kb)#taking the fixed point of the knowlwdge base
    #watch out because fp is a set of strings
    
    atoms = {c.head for c in kb.clauses if c.isAtom()}#it is sure that none of atoms is a disease
    #atoms = set(atoms)
    
    
    return list(fp.difference(atoms))
    
    


def get_prob_model(kb = KB()):
    """
    La funzione ritorna una lista di oggetti ProbItem che indicano un modello probabilistico che riporta la probabilità
    che il paziente abbia la malattia considerata.
    Ovviamente se non ci sono malattie non ha senso andare ad eseguire tutti i passi.
    Per cui, per ragioni di complessità, si considerano uscite anticipate.
    """
    disease = get_disease(kb)#prendo le malattie vere dalla base di conoscenza
    if len(disease)==0: return [] #non si avvera alcuna malattia
    model = []#istanzio la variabile modello che avrà il ranking delle probabilità associate ad ogni malattia
    for d in disease:
        #appendo un oggetto di tipo ProbItem 
        model.append(ProbItem(d,get_prob(d,kb)))
        
    sumProb = sum(item.prob for item in model )#calcolo la somma delle probabilità (serve per avere poi la somma pari a uno)
    for item in model :
        item.prob = item.prob/sumProb#vado a sostituire la probabilità aggiustata
    model.sort(reverse = True)  #ordino
    return model#restituisco
    
    
        
def get_prob(disease, kb=KB()):
    """
    Restituisce la probabilità che quella malattia sia "vera" 
    Pre-condizione-> la malattia deve avere una clausola associata , con un body diverso da []    
    """
    allSym = len({c.head for c in kb.clauses if c.isAtom()})#prendo il numero dei sintomi che sono veri
    for c in kb.clauses:#cerco la mia malattia
        if c.head == disease:
            return len(c.body)/allSym#restituisco il numero di sintomi che ha su tutti quelli veri    
        
"""
Function useful for testing
"""


import KB_Parser as parser
import openpyxl as xl
class Test:
    def __init__(self,file,row):
        """
        Constructor of a test object:
        file supported-> csv,xlsx
        row->which test do you want to do?
        """
        self.assumables = []
        
        if row <=0:#row not valid -> row 0 in an excel file does not exist
            raise ValueError("Cannot index the test at row "+str(row))
            
        if row ==1:#row 1 in formatted excel file is the row of the synthoms
            raise ValueError("Row 1 is not a disease")
        
        #checking file supported
        if (not parser.isCSV(file)) and (not parser.isExcel(file)):
            raise ValueError("file not supported-> must be csv or xlsx")
            
        #if necessary convert the file    
        if(parser.isCSV(file)):
            file= parser.csv_to_excel(file)
            
        #taking workbook
        wb = xl.load_workbook(file)
        #taking right sheet
        sheet = wb["Sheet1"]
        col = 2
        while sheet.cell(1,col).value != "prognosis":#for each synthom
            if sheet.cell(row,col).value== 1:#if that synthom is in that disease
                self.assumables.append(Clause(sheet.cell(1,col).value))#appending that synthom
            col+=1
        self.prognosis = sheet.cell(row,col).value#saving eight prognosis
        
        
        
    def verify(self,train_file):
        kb = parser.buildKB(train_file)#building the knowledge base
        kb.clauses+=self.assumables#adding test synthoms
        return self.prognosis in get_disease(kb)#if the prognosis is in the set of diseases is ok-> return True
        
        
    
    
    
def real_filename(filename= ""):
    return os.path.realpath(filename)
    
    
    
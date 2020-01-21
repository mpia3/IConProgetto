# -*- coding: utf-8 -*-

"""
This file implements all funtion nedded to parse the dataset, made up in a csv file (or in a xlsx file)
"""


import openpyxl as xl
import pandas as pd
from logicProblem import KB,Clause
import re





def isExcel(filename):
    "returns True if is an excel file, false otherwise"
    match= re.search(".xlsx$",filename)
    if match == None:
        return False
    return True



def isCSV(filename):
    match=re.search(".csv$",filename)
    if match==None:
        return False
    return True



def csv_to_excel(file):
    """
    Function that converts a csv file into a xlsx file.
    
    It creates a new file which name is file*_Excel.xlsx
    
    THROWS PermissionError
    
    """
    data = pd.read_csv(file)
    data.head()
    file =file.replace(".csv","_Excel.xlsx")
    try:    
        data.to_excel(file)
    except PermissionError:
        pass#il file è già stato creato
    return file#giving information to to caller
    
    

def parseSyn(file):
    """
    This function returns all synthoms took from the training file.
    
    It tries to convert the csv file into Excel file.
    Than it loops over all the synthoms and returns its list.
    
    warning-> file must be *.xlsx
    """    
    
    wb = xl.load_workbook(filename=file)
    sheet = wb["Sheet1"]
    #reading columns
    syn = []
    row = 1
    column = 2
    c = sheet.cell(row,column).value
    
    while c !="prognosis":
        syn.append(c)
        column+=1
        c = sheet.cell(row,column).value
    return syn


def buildKB(file):
    """
    This function returns a Knowledge Base with all clauses.
    It takes the file which contains the knowlwdge base and it converts that into an object KB
    
    warining-> file must be a csv or a xlsx file.
    raises-> ValueError if file is not supported
    """
    if (not isExcel(file)) and (not isCSV(file)):#control of file extension
        raise ValueError("The file passed in not supported (must be csv or xlsx)")
    
    
    if isCSV(file):#if necessary, make the conversion from csv to xlsx
        file = csv_to_excel(file)
        print(file)
    syn = parseSyn(file)#taking all synthoms from file
    
    wb = xl.load_workbook(filename=file)
    sheet = wb["Sheet1"]
    syn_len = len(syn)
    row = 2
    
    
    clauses = []#initialization of all clauses
    while True:#taking all clauses
        if sheet.cell(row,2).value == None:#if i am out of rows (I have finished all diseases)
            break
        
        #initialization of all synthoms for a specific disease
        nec_syn = [syn[col-2] for col in range (2,syn_len+2) if sheet.cell(row,col).value ==1]
        dis = sheet.cell(row,syn_len+2).value#saving disease name
        clauses.append(Clause(dis,nec_syn))#building and appending the entire clause
        row+=1
    return KB(clauses)
    
    

"Test"
def test(train,testing):
    """
    Test function which tries to give the correct disease from the given synthoms of the 1st row 
    (Testing.xlsx)
    
    The result set that is given contains all the synthoms true (taken from user) and all the diseases true
    tanks to the reasoner
    """
    
    KB = buildKB(train)
    
    wbTest = xl.load_workbook(filename = testing)#open test workbook
    sheet_test = wbTest["Sheet1"]
    col = 2
    assumables = []
    while sheet_test.cell(1,col).value != "prognosis":
        if sheet_test.cell(2,col).value==1:
            assumables.append(Clause(sheet_test.cell(1,col).value))
        col+=1
    
    
    KB.clauses+= assumables
    for c in KB.statements:
        print("Head: ",c.head)
        print("Body: ",c.body)
    
    
    from logicBottomUp import fixed_point
    fix_point = fixed_point(KB)
    print(fix_point)


